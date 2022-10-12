import openpyxl

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]
# wb.create_sheet("Sheet2", 0)
# wb.remove(sheet)
print(sheet.max_row)
print(sheet.max_column)

cell = sheet["a1"]
cell = sheet.cell(row=1, column=1)

print(cell.value)
print(cell.row)
print(cell.column)
print(cell.coordinate)

# Iterando por filas y columnas para acceder al valor
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)

# Otras formas de acceder a un rango de celdas
column = sheet["a"]
cells = sheet["a:c"]
cells = sheet["a1:c3"]
rows = sheet[1:3]

# Agregando data a la hoja
sheet.append([1, 2, 3])
wb.save("transactions2.xlsx")

# Principio de separacion comando-consulta
# Una consulta no deberia generar un cambio en el sistema
sheet.cell()  # Viola el principio porque crea la celda si no existe
